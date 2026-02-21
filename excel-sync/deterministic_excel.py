#!/usr/bin/env python3
"""
Deterministic Excel Updater for Hosting Deals

Replaces manual Power Query with automated, structured data updates.
- Fetches from deterministic API (let_api_fetcher.py)
- Programmatic deduplication (not formula-based)
- Auto-generates Offer Details sheet
- Creates Dashboard with real metrics
- Color-coded status (NEW=green, ACTIVE=blue, EXPIRED=gray)

Usage:
    python deterministic_excel.py --fetch
    python deterministic_excel.py --update data/deals.json
    python deterministic_excel.py --full-sync
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: pandas/openpyxl not installed. Run: pip install pandas openpyxl")

# Paths
DEFAULT_EXCEL = Path.home() / "workspace" / "Hosting-Deals-Tracker.xlsx"
DATA_DIR = Path.home() / ".let-automation" / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

if EXCEL_AVAILABLE:
    # Color scheme (matching Excel's theme)
    COLORS = {
        "NEW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),      # Light green
        "ACTIVE": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),    # Light blue
        "EXPIRED": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),   # Gray
        "DUPLICATE": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # Light red
        "HEADER": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),    # Dark blue
    }
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    FONT_BOLD = Font(bold=True)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
else:
    COLORS = {}
    FONT_HEADER = None
    FONT_BOLD = None
    ALIGN_CENTER = None


class DeterministicExcelManager:
    """Deterministic Excel manager with auto-generated sheets."""
    
    def __init__(self, excel_path: Path | None = None):
        self.excel_path = excel_path or DEFAULT_EXCEL
        self.deals_df: pd.DataFrame | None = None
        self.offer_details_df: pd.DataFrame | None = None
    
    def fetch_latest_deals(self) -> list[dict]:
        """Run the fetcher to get latest deals."""
        
        print("📡 Fetching latest deals...")
        
        # Import and run fetcher
        sys.path.insert(0, str(Path(__file__).parent.parent / "fetcher"))
        from let_api_fetcher import LETFetcher
        
        fetcher = LETFetcher()
        deals = fetcher.fetch_and_parse(pages=3, fetch_details=False)
        
        # Save for reference
        fetcher.save_deals(deals, "latest_deals.json")
        
        return deals
    
    def deduplicate(self, deals: list[dict]) -> tuple[list[dict], list[dict]]:
        """
        Deterministic deduplication based on thread_id + title.
        Returns (unique_deals, duplicate_deals).
        """
        
        seen = {}
        unique = []
        duplicates = []
        
        for deal in deals:
            # Create dedup key
            thread_id = deal.get("thread_id") or ""
            title = deal.get("title", "")[:50].lower()  # First 50 chars normalized
            
            key = f"{thread_id}:{title}"
            
            if key in seen:
                deal["_duplicate_of"] = seen[key]
                deal["_is_duplicate"] = True
                duplicates.append(deal)
            else:
                seen[key] = deal.get("url", "")
                deal["_is_duplicate"] = False
                unique.append(deal)
        
        print(f"📊 Deduplication: {len(unique)} unique, {len(duplicates)} duplicates")
        return unique, duplicates
    
    def create_deals_tracker(self, deals: list[dict]) -> pd.DataFrame:
        """Create the main Deals Tracker DataFrame."""
        
        # Normalize deals to flat structure
        rows = []
        for deal in deals:
            row = {
                # Core identification
                "Thread ID": deal.get("thread_id"),
                "Thread Title": deal.get("title", ""),
                "Author": deal.get("author", "Unknown"),
                "Post Date": deal.get("post_date", ""),
                
                # Categorization
                "Provider": deal.get("provider") or "Unknown",
                "Category": deal.get("category", "VPS"),
                "Status": deal.get("status", "UNKNOWN"),
                "Is Duplicate": "YES" if deal.get("_is_duplicate") else "NO",
                
                # Pricing
                "Price (Monthly)": deal.get("price_monthly"),
                "Price (Yearly)": deal.get("price_yearly"),
                "Billing": "Monthly" if deal.get("price_monthly") else "Yearly",
                
                # Technical Specs
                "vCPU": deal.get("cpu"),
                "RAM (GB)": deal.get("ram_gb"),
                "Storage (GB)": deal.get("storage_gb"),
                "Storage Type": deal.get("storage_type"),
                "Bandwidth (GB)": deal.get("bandwidth"),
                "IPv4 Count": deal.get("ipv4_count", 1),
                "IPv6": "YES" if deal.get("ipv6") else "NO",
                "Location": deal.get("location") or "Unknown",
                
                # Metadata
                "Source URL": deal.get("url", ""),
                "Fetch Date": deal.get("fetch_date", datetime.now().isoformat()),
                "Content Preview": deal.get("content_preview", "")[:200],
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Sort by date (newest first)
        df["Post Date"] = pd.to_datetime(df["Post Date"], errors="coerce")
        df = df.sort_values("Post Date", ascending=False)
        
        return df
    
    def create_offer_details(self, deals: list[dict]) -> pd.DataFrame:
        """Create the Offer Details sheet with technical specs."""
        
        # Filter to unique deals with specs
        spec_deals = [d for d in deals if not d.get("_is_duplicate")]
        
        rows = []
        for deal in spec_deals:
            # Only include deals with some specs
            if deal.get("cpu") or deal.get("ram_gb") or deal.get("storage_gb"):
                row = {
                    "Thread Title": deal.get("title", "")[:80],
                    "Provider": deal.get("provider") or "Unknown",
                    "Category": deal.get("category", "VPS"),
                    "Post Date": deal.get("post_date", ""),
                    "Status": deal.get("status", "UNKNOWN"),
                    
                    # Technical specs
                    "vCPU": deal.get("cpu"),
                    "RAM (GB)": deal.get("ram_gb"),
                    "Storage (GB)": deal.get("storage_gb"),
                    "Storage Type": deal.get("storage_type", "SSD"),
                    "Bandwidth (GB)": deal.get("bandwidth"),
                    "IPv4": deal.get("ipv4_count", 1),
                    "IPv6": "YES" if deal.get("ipv6") else "NO",
                    "Location": deal.get("location") or "Unknown",
                    
                    # Pricing
                    "Monthly Price": deal.get("price_monthly"),
                    "Yearly Price": deal.get("price_yearly"),
                    "Source URL": deal.get("url", ""),
                }
                rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Sort by provider and price
        df = df.sort_values(["Provider", "Monthly Price"])
        
        return df
    
    def create_dashboard_data(self, deals_df: pd.DataFrame) -> dict[str, Any]:
        """Generate dashboard metrics."""
        
        # Filter to unique deals only
        unique_df = deals_df[deals_df["Is Duplicate"] == "NO"]
        
        metrics = {
            "total_deals": len(deals_df),
            "unique_deals": len(unique_df),
            "duplicate_deals": len(deals_df) - len(unique_df),
            
            # By status
            "new_deals": len(unique_df[unique_df["Status"] == "NEW"]),
            "active_deals": len(unique_df[unique_df["Status"] == "ACTIVE"]),
            "expired_deals": len(unique_df[unique_df["Status"] == "EXPIRED"]),
            
            # By category
            "vps_deals": len(unique_df[unique_df["Category"] == "VPS"]),
            "dedicated_deals": len(unique_df[unique_df["Category"] == "Dedicated"]),
            "shared_deals": len(unique_df[unique_df["Category"] == "Shared"]),
            
            # By provider (top 10)
            "top_providers": unique_df["Provider"].value_counts().head(10).to_dict(),
            
            # Price stats
            "price_stats": {
                "min_monthly": unique_df["Price (Monthly)"].min(),
                "max_monthly": unique_df["Price (Monthly)"].max(),
                "avg_monthly": unique_df["Price (Monthly)"].mean(),
                "median_monthly": unique_df["Price (Monthly)"].median(),
            },
            
            # Specs averages
            "avg_specs": {
                "vcpu": unique_df["vCPU"].mean(),
                "ram": unique_df["RAM (GB)"].mean(),
                "storage": unique_df["Storage (GB)"].mean(),
            },
            
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        
        return metrics

    def _prepare_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert timezone-aware datetimes to timezone-naive for Excel compatibility."""
        prepared = df.copy()
        for col in prepared.columns:
            try:
                if pd.api.types.is_datetime64_any_dtype(prepared[col]):
                    tz = getattr(prepared[col].dt, "tz", None)
                    if tz is not None:
                        prepared[col] = prepared[col].dt.tz_convert(None)
            except Exception:
                continue
        return prepared
    
    def generate_excel(self, deals: list[dict] | None = None, 
                       backup: bool = True) -> Path:
        """Generate complete Excel file with all sheets."""
        
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas and openpyxl required")
        
        # Fetch deals if not provided
        if deals is None:
            deals = self.fetch_latest_deals()
        
        # Deduplicate
        unique_deals, duplicate_deals = self.deduplicate(deals)
        all_deals = unique_deals + duplicate_deals
        
        # Create dataframes
        print("\n📊 Creating worksheets...")
        deals_df = self.create_deals_tracker(all_deals)
        offer_df = self.create_offer_details(unique_deals)
        dashboard_metrics = self.create_dashboard_data(deals_df)
        
        # Backup existing file
        if backup and self.excel_path.exists():
            backup_path = self.excel_path.with_suffix(
                f".backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            )
            shutil.copy(self.excel_path, backup_path)
            print(f"💾 Backup created: {backup_path}")
        
        # Create Excel writer
        print(f"📝 Writing to: {self.excel_path}")
        
        with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
            
            # Sheet 1: Deals Tracker
            print("   Sheet 1: Deals Tracker")
            deals_df_xlsx = self._prepare_dataframe_for_excel(deals_df)
            deals_df_xlsx.to_excel(writer, sheet_name="Deals Tracker", index=False)
            self._format_deals_sheet(writer.sheets["Deals Tracker"], deals_df_xlsx)
            
            # Sheet 2: Deduplicated View (unique only)
            print("   Sheet 2: Deduplicated View")
            unique_df = deals_df_xlsx[deals_df_xlsx["Is Duplicate"] == "NO"]
            unique_df_xlsx = self._prepare_dataframe_for_excel(unique_df)
            unique_df_xlsx.to_excel(writer, sheet_name="Deduplicated View", index=False)
            self._format_deals_sheet(writer.sheets["Deduplicated View"], unique_df_xlsx)
            
            # Sheet 3: Offer Details
            print("   Sheet 3: Offer Details")
            if not offer_df.empty:
                offer_df_xlsx = self._prepare_dataframe_for_excel(offer_df)
                offer_df_xlsx.to_excel(writer, sheet_name="Offer Details", index=False)
                self._format_offer_sheet(writer.sheets["Offer Details"], offer_df_xlsx)
            
            # Sheet 4: Dashboard
            print("   Sheet 4: Dashboard")
            self._create_dashboard_sheet(writer, dashboard_metrics)
            
            # Sheet 5: Raw Data (for reference)
            print("   Sheet 5: Raw Data")
            raw_df = pd.DataFrame(all_deals)
            raw_df_xlsx = self._prepare_dataframe_for_excel(raw_df)
            raw_df_xlsx.to_excel(writer, sheet_name="Raw Data", index=False)
        
        print(f"\n✅ Excel file generated: {self.excel_path}")
        return self.excel_path
    
    def _format_deals_sheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting to deals sheet."""
        
        # Header formatting
        for cell in worksheet[1]:
            cell.fill = COLORS["HEADER"]
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Color-code status column
        status_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == "Status":
                status_col = idx
                break
        
        if status_col:
            for row in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row, column=status_col)
                status = status_cell.value
                if status in COLORS:
                    status_cell.fill = COLORS[status]
        
        # Mark duplicates
        dup_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == "Is Duplicate":
                dup_col = idx
                break
        
        if dup_col:
            for row in range(2, worksheet.max_row + 1):
                dup_cell = worksheet.cell(row=row, column=dup_col)
                if dup_cell.value == "YES":
                    # Highlight entire row
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = COLORS["DUPLICATE"]
        
        # Freeze header
        worksheet.freeze_panes = "A2"
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
    
    def _format_offer_sheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting to offer details sheet."""
        
        # Header formatting
        for cell in worksheet[1]:
            cell.fill = COLORS["HEADER"]
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
        
        # Auto-adjust widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Color-code status
        status_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == "Status":
                status_col = idx
                break
        
        if status_col:
            for row in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row, column=status_col)
                status = status_cell.value
                if status in COLORS:
                    status_cell.fill = COLORS[status]
        
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    
    def _create_dashboard_sheet(self, writer, metrics: dict):
        """Create dashboard with metrics."""
        
        # Create dataframe for dashboard
        dashboard_data = [
            ["METRIC", "VALUE", "NOTES"],
            ["", "", ""],
            ["SUMMARY", "", ""],
            ["Total Deals", metrics["total_deals"], "All fetched deals"],
            ["Unique Deals", metrics["unique_deals"], "After deduplication"],
            ["Duplicate Deals", metrics["duplicate_deals"], "Duplicates removed"],
            ["", "", ""],
            ["BY STATUS", "", ""],
            ["NEW Deals (< 7 days)", metrics["new_deals"], "Green highlight"],
            ["ACTIVE Deals (7-30 days)", metrics["active_deals"], "Blue highlight"],
            ["EXPIRED Deals (> 30 days)", metrics["expired_deals"], "Gray highlight"],
            ["", "", ""],
            ["BY CATEGORY", "", ""],
            ["VPS Deals", metrics["vps_deals"], ""],
            ["Dedicated Deals", metrics["dedicated_deals"], ""],
            ["Shared Hosting", metrics["shared_deals"], ""],
            ["", "", ""],
            ["PRICING STATS", "", ""],
            ["Min Monthly Price", f"${metrics['price_stats']['min_monthly']:.2f}", ""],
            ["Max Monthly Price", f"${metrics['price_stats']['max_monthly']:.2f}", ""],
            ["Avg Monthly Price", f"${metrics['price_stats']['avg_monthly']:.2f}", ""],
            ["Median Monthly Price", f"${metrics['price_stats']['median_monthly']:.2f}", ""],
            ["", "", ""],
            ["AVERAGE SPECS", "", ""],
            ["Avg vCPU", f"{metrics['avg_specs']['vcpu']:.1f}", ""],
            ["Avg RAM (GB)", f"{metrics['avg_specs']['ram']:.1f}", ""],
            ["Avg Storage (GB)", f"{metrics['avg_specs']['storage']:.1f}", ""],
            ["", "", ""],
            ["TOP PROVIDERS", "", ""],
        ]
        
        # Add top providers
        for provider, count in list(metrics["top_providers"].items())[:5]:
            dashboard_data.append([provider, count, ""])
        
        dashboard_data.append(["", "", ""])
        dashboard_data.append(["Last Updated", metrics["last_updated"], ""])
        
        df = pd.DataFrame(dashboard_data)
        df.to_excel(writer, sheet_name="Dashboard", index=False, header=False)
        
        # Format dashboard
        worksheet = writer.sheets["Dashboard"]
        
        # Section headers
        for row in [3, 9, 15, 18, 24]:
            for col in range(1, 4):
                cell = worksheet.cell(row=row, column=col)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Adjust widths
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 20
        worksheet.column_dimensions["C"].width = 40


def main():
    parser = argparse.ArgumentParser(description="Deterministic Excel updater")
    parser.add_argument("--file", "-f", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--fetch", action="store_true", help="Fetch latest deals first")
    parser.add_argument("--update", "-u", type=Path, help="Update from JSON file")
    parser.add_argument("--full-sync", action="store_true", help="Fetch + update")
    parser.add_argument("--no-backup", action="store_true", help="Skip backup")
    
    args = parser.parse_args()
    
    if not EXCEL_AVAILABLE:
        print("❌ pandas and openpyxl required. Run: pip install pandas openpyxl")
        return 1
    
    manager = DeterministicExcelManager(args.file)
    
    deals = None
    
    if args.full_sync or args.fetch:
        deals = manager.fetch_latest_deals()
    
    if args.update:
        deals = json.loads(args.update.read_text())
    
    # Generate Excel
    manager.generate_excel(deals=deals, backup=not args.no_backup)
    
    print("\n✅ Done!")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
